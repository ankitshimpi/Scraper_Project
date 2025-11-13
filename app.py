from flask import Flask, render_template, request, redirect, url_for, session, send_file, flash
import os
import re
import sqlite3
import pandas as pd

# typing for broader Python compat
from typing import List, Set

# NEW: for EXE-friendly launch and resource paths
import threading
import webbrowser
import sys
from waitress import serve  # make sure 'waitress' is installed when building

# >>> ADDED: styling for Excel highlights
from openpyxl.styles import PatternFill, Font
# <<< ADDED

# >>> ADDED: for moving files on finalize
import shutil
# <<< ADDED

# scrapers
from scraper_products import run_products_scraper
from scraper_competitor import run_competitor_scraper

# ---------- Helper so PyInstaller can find templates/static ----------
def resource_path(rel_path: str) -> str:
    """
    Resolve paths for bundled resources both in dev and in PyInstaller EXE.
    """
    if hasattr(sys, "_MEIPASS"):
        base = sys._MEIPASS  # type: ignore[attr-defined]
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, rel_path)

# Use explicit template/static folders so Flask finds them in the EXE
app = Flask(
    __name__,
    template_folder=resource_path("templates"),
    static_folder=resource_path("static"),
)
app.secret_key = "supersecretkey"  # Change this in production
DB_FILE = "users.db"

def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL
        )
    """)
    conn.commit()
    conn.close()

init_db()

# >>> ADDED: utilities for normalization, ordering, writing, cleaning, and staging
def _norm_asin(x: str) -> str:
    if pd.isna(x):
        return ""
    x = str(x).strip().upper()
    return x if re.fullmatch(r"B0[A-Z0-9]{8}", x) else ""

def _ensure_first_column(df: pd.DataFrame, first_col: str) -> pd.DataFrame:
    df = df.copy()
    if first_col not in df.columns:
        df[first_col] = pd.NA
    cols = [first_col] + [c for c in df.columns if c != first_col]
    return df[cols]

def _build_search_map_from_comp_df(comp_df: pd.DataFrame) -> dict:
    """
    Map Competitor_ASIN -> Search_ASIN from Helium10 'Competitors' sheet.
    """
    if "Competitor_ASIN" not in comp_df.columns or "Search_ASIN" not in comp_df.columns:
        return {}
    df = comp_df[["Competitor_ASIN", "Search_ASIN"]].copy()
    df["Competitor_ASIN"] = df["Competitor_ASIN"].map(_norm_asin)
    df["Search_ASIN"] = df["Search_ASIN"].map(_norm_asin)
    df = df[(df["Competitor_ASIN"] != "") & (df["Search_ASIN"] != "")]
    return dict(zip(df["Competitor_ASIN"], df["Search_ASIN"]))

def _add_search_asin_to_amazon_frames(
    df_products: pd.DataFrame,
    df_sellers: pd.DataFrame,
    search_map: dict,
    seed_asins: list
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Returns new (products, sellers) DataFrames with 'Search_ASIN' populated and moved to first column.
    Rule:
      - If ASIN maps via search_map, use that.
      - Else if ASIN is one of the seeds, Search_ASIN = ASIN (self).
      - Else NaN.
    """
    seed_set = set(seed_asins)

    def _attach(df: pd.DataFrame) -> pd.DataFrame:
        if df is None:
            return df
        df = df.copy()
        if "ASIN" in df.columns:
            asins = df["ASIN"].astype(str).map(_norm_asin)
            mapped = asins.map(search_map)
            df["Search_ASIN"] = [
                (m if pd.notna(m) and str(m) != "" else (a if a in seed_set else pd.NA))
                for m, a in zip(mapped, asins)
            ]
        else:
            df["Search_ASIN"] = pd.NA
        df = _ensure_first_column(df, "Search_ASIN")
        return df

    return _attach(df_products), _attach(df_sellers)

def _resolve_out_path(base_path_from_input_file: str, out_path: str) -> str:
    """
    If out_path is absolute, use it directly; otherwise write next to base file.
    """
    if os.path.isabs(out_path):
        return out_path
    return os.path.join(os.path.dirname(base_path_from_input_file), out_path)

def _write_helium10_by_asin(competitor_file_path: str, seed_asins: list, out_path: str) -> str:
    """
    Write {file_name}_competitors_by_asin.xlsx
    One sheet per input ASIN (sheet name = ASIN), filtered by 'Search_ASIN'.
    'Search_ASIN' is the first column.
    """
    xl = pd.ExcelFile(competitor_file_path)
    if "Competitors" not in xl.sheet_names:
        return ""
    comp_df = pd.read_excel(competitor_file_path, sheet_name="Competitors")
    if "Search_ASIN" not in comp_df.columns:
        return ""

    out_file = _resolve_out_path(competitor_file_path, out_path)
    os.makedirs(os.path.dirname(out_file), exist_ok=True)

    with pd.ExcelWriter(out_file, engine="openpyxl", mode="w") as writer:
        for asin in seed_asins:
            df_sub = comp_df[comp_df["Search_ASIN"].astype(str).map(_norm_asin) == asin].copy()
            if df_sub.empty:
                df_sub = comp_df.head(0).copy()
            df_sub = _ensure_first_column(df_sub, "Search_ASIN")
            df_sub.to_excel(writer, sheet_name=asin, index=False)
    return out_file

def _write_amazon_by_asin_with_sellers(
    amazon_file_path: str,
    seed_asins: list,
    search_map: dict,
    out_path: str
) -> str:
    """
    Write {file_name}_products_by_asin.xlsx
    For each input ASIN:
      - sheet '<ASIN>' with Products rows where Search_ASIN == ASIN (Search_ASIN first col)
      - sheet 'OtherSeller_<ASIN>' with OtherSellers rows where Search_ASIN == ASIN (Search_ASIN first col)
    Also highlight in each '<ASIN>' sheet the rows where ASIN == Search_ASIN (the input ASIN rows).
    """
    xl = pd.ExcelFile(amazon_file_path)
    df_products = pd.read_excel(amazon_file_path, sheet_name="Products") if "Products" in xl.sheet_names else pd.DataFrame()
    df_sellers  = pd.read_excel(amazon_file_path, sheet_name="OtherSellers") if "OtherSellers" in xl.sheet_names else pd.DataFrame()

    # Attach Search_ASIN and move to first column
    df_products, df_sellers = _add_search_asin_to_amazon_frames(df_products, df_sellers, search_map, seed_asins)

    out_file = _resolve_out_path(amazon_file_path, out_path)
    os.makedirs(os.path.dirname(out_file), exist_ok=True)

    with pd.ExcelWriter(out_file, engine="openpyxl", mode="w") as writer:
        # Write Products sheets per ASIN
        for asin in seed_asins:
            prod_sub = df_products[df_products["Search_ASIN"].astype(str).map(_norm_asin) == asin].copy()
            if prod_sub.empty:
                prod_sub = df_products.head(0).copy()
            prod_sub.to_excel(writer, sheet_name=asin, index=False)

        # Write OtherSellers sheets per ASIN
        for asin in seed_asins:
            sell_sub = df_sellers[df_sellers["Search_ASIN"].astype(str).map(_norm_asin) == asin].copy()
            if sell_sub.empty:
                sell_sub = df_sellers.head(0).copy()
            sell_sub.to_excel(writer, sheet_name=f"OtherSeller_{asin}", index=False)

        # === Highlight Products rows where ASIN == Search_ASIN on each '<ASIN>' sheet ===
        book = writer.book
        yellow = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # soft yellow
        bold_font = Font(bold=True)
        for asin in seed_asins:
            sheet_name = asin
            if sheet_name not in book.sheetnames:
                continue
            ws = book[sheet_name]

            # Find column indexes for ASIN and Search_ASIN
            headers = [cell.value for cell in ws[1]]
            try:
                col_asin = headers.index("ASIN") + 1
            except ValueError:
                col_asin = None
            try:
                col_search = headers.index("Search_ASIN") + 1
            except ValueError:
                col_search = None

            if not col_asin or not col_search:
                continue

            # Iterate data rows
            for r in range(2, ws.max_row + 1):
                val_asin = ws.cell(row=r, column=col_asin).value
                val_search = ws.cell(row=r, column=col_search).value
                if val_asin and val_search and str(val_asin).strip().upper() == str(val_search).strip().upper():
                    # highlight entire row
                    for c in range(1, ws.max_column + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.fill = yellow
                        cell.font = bold_font

    return out_file

def _cleanup_files(paths: List[str], keep: Set[str]):
    """
    Delete files in 'paths' that are not in 'keep', if they exist.
    """
    for p in paths:
        try:
            if p and os.path.exists(p) and p not in keep:
                os.remove(p)
        except Exception:
            # best-effort cleanup; ignore failures
            pass

def _make_staging_dir(save_path: str) -> str:
    """
    Create a hidden staging directory inside the user-provided input folder.
    Files will be moved to parent folder only when user clicks Download.
    """
    staging_dir = os.path.join(save_path, "_staging")
    os.makedirs(staging_dir, exist_ok=True)
    return staging_dir

def _is_in_staging(p: str) -> bool:
    if not p:
        return False
    parts = os.path.normpath(p).split(os.sep)
    return "_staging" in parts
# <<< ADDED


# >>> NEW HELPER FOR AD REPORT → PRODUCTS FLOW
def extract_unique_asins_from_ad_report(file_path: str) -> list:
    """
    1. Open Excel file at file_path
    2. Read sheet 'Sponsored Products Campaigns'
    3. Read sheet 'Sponsored Display Campaigns'
    4. From each, take column 'ASIN (Informational only)'
    5. Clean, dedupe, return list of ASIN strings (uppercased, trimmed)
    """
    target_sheets = [
        "Sponsored Products Campaigns",
        "Sponsored Display Campaigns",
    ]
    asin_col = "ASIN (Informational only)"
    asin_values = []

    try:
        xls = pd.ExcelFile(file_path)
    except Exception as e:
        print(f"[AdReport] Error opening file: {e}")
        return []

    for sheet in target_sheets:
        if sheet not in xls.sheet_names:
            print(f"[AdReport] Warning: sheet '{sheet}' not found. Skipping.")
            continue

        try:
            df = pd.read_excel(xls, sheet_name=sheet)
        except Exception as e:
            print(f"[AdReport] Error reading sheet '{sheet}': {e}")
            continue

        if asin_col not in df.columns:
            print(f"[AdReport] Warning: column '{asin_col}' not in '{sheet}'. Skipping.")
            continue

        col_series = (
            df[asin_col]
            .dropna()
            .astype(str)
            .str.strip()
            .str.upper()
        )

        asin_values.extend([val for val in col_series.tolist() if val])

    # dedupe preserving first-seen order
    unique = []
    seen = set()
    for a in asin_values:
        if a not in seen:
            seen.add(a)
            unique.append(a)

    return unique
# <<< NEW HELPER END


@app.route("/")
def home():
    if "user" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))

# --------- Register ---------
@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        try:
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("INSERT INTO users (username, password) VALUES (?, ?)", (username, password))
            conn.commit()
            conn.close()
            flash("Account created successfully! Please login.", "success")
            return redirect(url_for("login"))
        except sqlite3.IntegrityError:
            flash("Username already exists!", "danger")
            return redirect(url_for("register"))
    return render_template("register.html")

# --------- Login ---------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT * FROM users WHERE username=? AND password=?", (username, password))
        user = c.fetchone()
        conn.close()
        if user:
            session["user"] = username
            return redirect(url_for("dashboard"))
        else:
            flash("Invalid username or password", "danger")
            return redirect(url_for("login"))
    return render_template("login.html")

# --------- Logout ---------
@app.route("/logout")
def logout():
    session.pop("user", None)
    flash("Logged out successfully!", "info")
    return redirect(url_for("login"))

# --------- Landing Dashboard (now with 3 buttons/cards) ---------
@app.route("/dashboard", methods=["GET"])
def dashboard():
    if "user" not in session:
        return redirect(url_for("login"))
    # Renders dashboard.html (updated to include Ad Report → Products card)
    return render_template("dashboard.html")

# --------- Dedicated pages (separated dashboards) ---------
@app.route("/dashboard/amazon-products", methods=["GET"])
def amazon_products_page():
    if "user" not in session:
        return redirect(url_for("login"))
    return render_template("amazon_products.html", tables={}, saved_file=None)

@app.route("/dashboard/competitor", methods=["GET"])
def competitor_page():
    if "user" not in session:
        return redirect(url_for("login"))
    return render_template("competitor_auto.html", tables={}, saved_file=None)

# >>> NEW PAGE: Ad Report → Products
@app.route("/dashboard/adreport", methods=["GET"])
def adreport_page():
    if "user" not in session:
        return redirect(url_for("login"))
    # initially empty preview, like others
    return render_template("adreport_products.html", tables={}, saved_file=None)
# <<< NEW PAGE


# --------- Run Amazon Products scraper (manual mode) ---------
@app.route("/run/products", methods=["POST"])
def run_products():
    if "user" not in session:
        return redirect(url_for("login"))

    asin_input = request.form.get("asin_input", "")
    save_path  = request.form.get("save_path", "").strip()
    file_name  = request.form.get("file_name", "").strip()
    # NEW: optional marketplace from the manual form
    marketplace = request.form.get("marketplace", "").strip() or None

    if not asin_input or not save_path or not file_name:
        flash("Please fill all fields for Amazon Products.", "danger")
        return redirect(url_for("amazon_products_page"))

    if not os.path.exists(save_path):
        os.makedirs(save_path, exist_ok=True)

    tables = {}
    saved_file = None
    try:
        # Pass marketplace through (None -> auto/try all)
        saved_file = run_products_scraper(
            asin_input,
            save_path,
            file_name,
            marketplace=marketplace
        )

        # Render Amazon tables
        excel_file = pd.ExcelFile(saved_file)
        for sheet in excel_file.sheet_names:
            df = pd.read_excel(saved_file, sheet_name=sheet)
            tables[f"Amazon - {sheet}"] = df.to_html(
                classes="table table-striped table-bordered datatable",
                index=False,
                table_id=f"table_amz_{sheet.replace(' ', '_')}"
            )

        if marketplace:
            flash(f"Amazon Products completed for marketplace '{marketplace}'. File: {saved_file}", "success")
        else:
            flash(f"Amazon Products completed (auto marketplace). File: {saved_file}", "success")
    except Exception as e:
        flash(f"Error running Amazon Products scraper: {str(e)}", "danger")

    return render_template("amazon_products.html", tables=tables, saved_file=saved_file)

# --------- Run Competitor scraper (Helium10) -> then Amazon automatically -> produce ONLY two final files ---------
@app.route("/run/competitor", methods=["POST"])
def run_competitor():
    if "user" not in session:
        return redirect(url_for("login"))

    # Inputs from the form
    seed_asins_raw = request.form.get("asin_input", "")
    marketplace_input = request.form.get("marketplace", "").strip()
    save_path  = request.form.get("save_path", "").strip()
    file_name  = request.form.get("file_name", "").strip()

    if not (seed_asins_raw and marketplace_input and save_path and file_name):
        flash("Please fill all fields for Competitor scraper.", "danger")
        return redirect(url_for("competitor_page"))

    if not os.path.exists(save_path):
        os.makedirs(save_path, exist_ok=True)

    # >>> ADDED: create staging dir inside the provided input folder
    staging_dir = _make_staging_dir(save_path)

    tables = {}
    competitor_file = None
    amazon_file = None

    try:
        # 1) Run Helium10 and save competitor Excel
        competitor_file = run_competitor_scraper(
            asin_input=seed_asins_raw,
            marketplace_input_raw=marketplace_input,
            save_path=staging_dir,   # <<< write raw H10 file into staging
            file_name=file_name,
            headless=False  # Changed to True to avoid GUI issues and potential timeouts
        )

        # 2) Parse seeds and build union list (seed ∪ competitors) for Amazon
        seed_asins = [a.strip().upper() for a in seed_asins_raw.replace(",", " ").split() if a.strip()]
        seed_asins = [a for a in seed_asins if re.fullmatch(r"B0[A-Z0-9]{8}", a)]

        comp_df = pd.read_excel(competitor_file, sheet_name="Competitors")
        if "Competitor_ASIN" not in comp_df.columns:
            raise RuntimeError("The 'Competitors' sheet is missing the 'Competitor_ASIN' column.")

        comp_asins = (
            comp_df["Competitor_ASIN"]
            .dropna().astype(str).str.upper().str.strip().tolist()
        )
        comp_asins = [a for a in comp_asins if re.fullmatch(r"B0[A-Z0-9]{8}", a)]
        asin_union = sorted(set(seed_asins) | set(comp_asins))

        if not asin_union:
            flash("No valid ASINs found to run Amazon after Check(check format B0XXXXXXXX).", "warning")
            # Show Competitors (raw) so user can see something
            comp_xl = pd.ExcelFile(competitor_file)
            for sheet in comp_xl.sheet_names:
                df = pd.read_excel(competitor_file, sheet_name=sheet)
                tables[f"Competitor - {sheet}"] = df.to_html(
                    classes="table table-striped table-bordered datatable",
                    index=False,
                    table_id=f"table_comp_{sheet.replace(' ', '_')}"
                )
            saved_file = competitor_file
            flash(f"(Staged) Competitor file saved at: {competitor_file}", "info")
            return render_template("competitor_auto.html", tables=tables, saved_file=saved_file)

        # 3) Run Amazon on seed + competitors
        asin_for_amazon = " ".join(asin_union)
        amazon_file_name = f"{file_name}_products"
        amazon_file = run_products_scraper(
            asin_input=asin_for_amazon,
            save_path=staging_dir,       # <<< write raw Amazon file into staging
            file_name=amazon_file_name,
            marketplace=marketplace_input
        )

        # >>> CHANGED: STOP copying Amazon sheets into the Helium10 workbook (to avoid extra files)

        # >>> ADDED: Build final two files ONLY — but in STAGING
        # Mappings for Search_ASIN
        search_map = _build_search_map_from_comp_df(comp_df)

        # Final staged file absolute paths
        comp_by_asin_staged = os.path.join(staging_dir, f"{file_name}_competitors_by_asin.xlsx")
        amz_by_asin_staged  = os.path.join(staging_dir, f"{file_name}_products_by_asin.xlsx")

        # 4) Create Helium10-by-ASIN workbook (Search_ASIN first col) in staging
        comp_by_asin_staged = _write_helium10_by_asin(
            competitor_file_path=competitor_file,
            seed_asins=seed_asins,
            out_path=comp_by_asin_staged  # absolute
        )
        if comp_by_asin_staged:
            flash(f"(Staged) Helium10 (by ASIN): {comp_by_asin_staged}", "success")

        # 5) Create Amazon-by-ASIN workbook in staging with separate OtherSellers sheets
        amz_by_asin_staged = _write_amazon_by_asin_with_sellers(
            amazon_file_path=amazon_file,
            seed_asins=seed_asins,
            search_map=search_map,
            out_path=amz_by_asin_staged  # absolute
        )
        if amz_by_asin_staged:
            flash(f"(Staged) Amazon (by ASIN): {amz_by_asin_staged}", "success")

        # >>> ADDED: Clean up intermediates so ONLY the two staged final files remain in staging
        _cleanup_files(
            paths=[competitor_file, amazon_file],
            keep={comp_by_asin_staged, amz_by_asin_staged}
        )

        # >>> ADDED: Save staging info in session; finalization happens on /download click
        session["staging_info"] = {
            "final_dir": save_path,
            "files": [comp_by_asin_staged, amz_by_asin_staged]
        }

        # 6) For display: read from the two staged final files
        # Helium10 by-ASIN tabs
        if comp_by_asin_staged and os.path.exists(comp_by_asin_staged):
            comp_xl = pd.ExcelFile(comp_by_asin_staged)
            for sheet in comp_xl.sheet_names:
                df = pd.read_excel(comp_by_asin_staged, sheet_name=sheet)
                tables[f"H10 - {sheet}"] = df.to_html(
                    classes="table table-striped table-bordered datatable",
                    index=False,
                    table_id=f"table_h10_{sheet.replace(' ', '_')}"
                )
        # Amazon by-ASIN tabs (Products and OtherSeller_*)
        if amz_by_asin_staged and os.path.exists(amz_by_asin_staged):
            amz_xl = pd.ExcelFile(amz_by_asin_staged)
            for sheet in amz_xl.sheet_names:
                df = pd.read_excel(amz_by_asin_staged, sheet_name=sheet)
                tables[f"Amazon - {sheet}"] = df.to_html(
                    classes="table table-striped table-bordered datatable",
                    index=False,
                    table_id=f"table_amz_{sheet.replace(' ', '_')}"
                )

        # Key success messages
        flash(f"Competitor scraping completed (marketplace '{marketplace_input}'). Files are staged.", "info")
        flash("Click 'Download Excel file' to finalize → files will be moved to your input folder path.", "warning")

        # IMPORTANT: Set the main download button to point at ONE staged file.
        # Clicking it will finalize BOTH files into the input folder and also download this file.
        saved_file = comp_by_asin_staged

    except Exception as e:
        flash(f"Error in Competitor → Amazon pipeline: {str(e)}", "danger")
        # Try to at least surface one file if available
        saved_file = (locals().get("comp_by_asin_staged") or
                      locals().get("competitor_file") or
                      locals().get("amz_by_asin_staged") or
                      locals().get("amazon_file"))

    return render_template("competitor_auto.html", tables=tables, saved_file=saved_file)

# >>> NEW: Run Ad Report → Products
@app.route("/run/adreport", methods=["POST"])
def run_adreport():
    """
    Flow:
    1. Get inputs:
        - file_path (Excel ad report)
        - marketplace
        - save_path
        - file_name
    2. Extract ASINs from:
        'Sponsored Products Campaigns' / 'Sponsored Display Campaigns'
        column 'ASIN (Informational only)'
    3. De-dupe ASINs
    4. Call run_products_scraper with those ASINs
    5. Render preview tables using adreport_products.html
    """
    if "user" not in session:
        return redirect(url_for("login"))

    file_path = request.form.get("file_path", "").strip()
    marketplace_input = request.form.get("marketplace", "").strip() or None
    save_path = request.form.get("save_path", "").strip()
    file_name = request.form.get("file_name", "").strip()

    if not file_path or not save_path or not file_name:
        flash("Please fill all fields for Ad Report → Products.", "danger")
        return redirect(url_for("adreport_page"))

    if not os.path.exists(save_path):
        os.makedirs(save_path, exist_ok=True)

    # Extract unique ASINs from ad report
    asin_list = extract_unique_asins_from_ad_report(file_path)

    if not asin_list:
        flash("No ASINs found in the provided report.", "warning")
        return render_template("adreport_products.html", tables={}, saved_file=None)

    asin_input_for_products = " ".join(asin_list)

    tables = {}
    saved_file = None

    try:
        # Reuse existing products scraper with these ASINs
        saved_file = run_products_scraper(
            asin_input=asin_input_for_products,
            save_path=save_path,
            file_name=file_name,
            marketplace=marketplace_input
        )

        # Build preview tables from the resulting Excel
        excel_file = pd.ExcelFile(saved_file)
        for sheet in excel_file.sheet_names:
            df = pd.read_excel(saved_file, sheet_name=sheet)
            tables[f"Amazon - {sheet}"] = df.to_html(
                classes="table table-striped table-bordered datatable",
                index=False,
                table_id=f"table_adrep_{sheet.replace(' ', '_')}"
            )

        flash("Ad Report → Products completed successfully.", "success")

    except Exception as e:
        flash(f"Error running Ad Report → Products: {str(e)}", "danger")

    return render_template("adreport_products.html", tables=tables, saved_file=saved_file)
# <<< NEW ROUTE END


# --------- Download File (also finalizes: move staged → input folder path on click) ---------
@app.route("/download")
def download():
    file_path = request.args.get("file")

    if not file_path or not os.path.exists(file_path):
        flash("File not found!", "danger")
        return redirect(url_for("dashboard"))

    # If the requested path is inside staging, first finalize:
    if _is_in_staging(file_path):
        staging_info = session.get("staging_info")
        if not staging_info:
            flash("Staging state not found; cannot finalize.", "danger")
            return redirect(url_for("dashboard"))

        final_dir = staging_info.get("final_dir")
        files = staging_info.get("files", [])
        if not final_dir or not files:
            flash("Invalid staging info; cannot finalize.", "danger")
            return redirect(url_for("dashboard"))

        # Move all staged files into the final input folder path
        moved_any = False
        for staged in files:
            if os.path.exists(staged):
                dest = os.path.join(final_dir, os.path.basename(staged))
                try:
                    # If destination exists, replace it
                    if os.path.exists(dest):
                        os.remove(dest)
                    shutil.move(staged, dest)
                    moved_any = True
                except Exception as e:
                    flash(f"Could not move {os.path.basename(staged)}: {e}", "warning")
            else:
                flash(f"Missing staged file: {staged}", "warning")

        # Try to remove staging dir if empty
        try:
            staging_dir = os.path.dirname(files[0])
            if os.path.isdir(staging_dir) and not os.listdir(staging_dir):
                os.rmdir(staging_dir)
        except Exception:
            pass

        if moved_any:
            flash("Files finalized into your input folder path.", "success")
        # Clear staging info
        session.pop("staging_info", None)

        # Update file_path to the finalized location of the clicked file so the browser still downloads it
        file_path = os.path.join(final_dir, os.path.basename(file_path))

    # Normal file download (either staged already finalized above, or a regular path)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)

    flash("File not found after finalize.", "danger")
    return redirect(url_for("dashboard"))

# ====== Launch: open browser and keep server alive in EXE ======
def _open_browser(port: int):
    # Open the login page in the default browser
    webbrowser.open_new(f"http://127.0.0.1:{port}/login")

if __name__ == "__main__":
    PORT = int(os.environ.get("PORT", 5000))
    threading.Timer(1.0, _open_browser, args=(PORT,)).start()
    serve(app, host="127.0.0.1", port=PORT)
